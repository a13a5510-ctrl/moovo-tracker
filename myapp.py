import asyncio
import os
import requests
import json
import pandas as pd
from datetime import datetime, timedelta
from playwright.async_api import async_playwright
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# [System] 任務日誌
def spy_log(msg): print(f"{msg}", flush=True)

# 🎯 台北時間 (UTC+8) 校正
now = datetime.utcnow() + timedelta(hours=8)
current_time_tp = now.strftime("%Y-%m-%d %H:%M")
file_date = now.strftime("%Y%m%d_%H%M")

# ================= 1. 配置 =================
LINE_TOKEN = os.getenv("LINE_CHANNEL_ACCESS_TOKEN")
LINE_TARGET = os.getenv("LINE_USER_ID")
AI_KEY = os.getenv("GEMINI_API_KEY")
DATA_FILE = "last_data.json"
EXCEL_FILE = f"Moovo_Report_{file_date}.xlsx"

# ================= 2. 功能函數 =================

def send_line(msg, file_url=None):
    """傳送 LINE 訊息，包含文字備援連結與檔案掛載"""
    url = "https://api.line.me/v2/bot/message/push"
    headers = {"Content-Type": "application/json", "Authorization": f"Bearer {LINE_TOKEN}"}
    
    # 📝 強化：將下載連結直接寫在文字訊息中，確保 100% 可讀取
    full_text = msg
    if file_url:
        full_text += f"\n\n📂 完整情報表下載：\n{file_url}"
    
    messages = [{"type": "text", "text": full_text}]
    
    # 同時嘗試發送檔案圖示消息
    if file_url:
        messages.append({
            "type": "file",
            "fileName": EXCEL_FILE,
            "originalContentUrl": file_url
        })
        
    payload = {"to": LINE_TARGET, "messages": messages}
    try:
        res = requests.post(url, headers=headers, json=payload, timeout=20)
        spy_log(f"[Line] 總傳送狀態碼: {res.status_code}")
    except Exception as e:
        spy_log(f"[Line] 傳輸崩潰: {e}")

def upload_excel(file_path):
    """三重上傳路徑：Catbox -> file.io -> transfer.sh"""
    if not os.path.exists(file_path):
        spy_log(f"[System] 錯誤：找不到檔案 {file_path}"); return None
    
    spy_log(f"[System] 準備上傳情報檔 ({os.path.getsize(file_path)} bytes)...")

    # 🚀 路徑 A: Catbox (穩定首選)
    try:
        spy_log("[System] 嘗試路徑 A (Catbox)...")
        with open(file_path, 'rb') as f:
            res = requests.post('https://catbox.moe/user/api.php', 
                                data={'reqtype': 'fileupload'}, 
                                files={'fileToUpload': f}, timeout=25)
            if res.status_code == 200 and "https" in res.text:
                return res.text.strip()
    except: pass

    # 🚀 路徑 B: file.io (第二備援)
    try:
        spy_log("[System] 嘗試路徑 B (file.io)...")
        with open(file_path, 'rb') as f:
            res = requests.post('https://file.io', files={'file': f}, timeout=20)
            if res.status_code == 200:
                data = res.json()
                if data.get('success'): return data.get('link')
    except: pass

    # 🚀 路徑 C: transfer.sh (終極備援)
    try:
        spy_log("[System] 嘗試路徑 C (transfer.sh)...")
        with open(file_path, 'rb') as f:
            res = requests.put(f'https://transfer.sh/{EXCEL_FILE}', data=f, timeout=20)
            if res.status_code == 200: return res.text.strip()
    except: pass

    spy_log("[System] ❌ 所有傳輸路徑均失敗")
    return None

def create_beautified_excel(changed, unchanged, filename):
    """建立並美化 Excel 報表"""
    spy_log("[System] 正在製作數位化美化情報表...")
    data = []
    for i in changed:
        data.append({"狀態": "⚡ 變動", "場站名稱": i['name'], "現有台數": i['curr'], "變動量": i['diff']})
    for i in unchanged:
        data.append({"狀態": "⚪ 穩定", "場站名稱": i['name'], "現有台數": i['curr'], "變動量": "0"})
    
    df = pd.DataFrame(data)
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Moovo即時情報')
        ws = writer.sheets['Moovo即時情報']
        
        # 🎨 設定樣式
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # 標題欄
        for cell in ws[1]:
            cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, center_align, border

        # 內容斑馬紋與格線
        for row in range(2, len(data) + 2):
            fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") if "⚡" in str(ws.cell(row=row, column=1).value) else None
            for cell in ws[row]:
                if fill: cell.fill = fill
                cell.alignment, cell.border = center_align, border

        # 自動調整欄寬
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 25
            
    return filename

def call_gemini(prompt):
    """呼叫精準的 Gemini 2.5-Flash 節點"""
    api_key = AI_KEY.strip()
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
    try:
        res = requests.post(url, headers={"Content-Type": "application/json"}, 
                            json={"contents": [{"parts": [{"text": prompt}]}]}, timeout=30)
        if res.status_code == 200:
            return res.json()['candidates'][0]['content']['parts'][0]['text']
        spy_log(f"[AI] 呼叫失敗，代碼：{res.status_code}")
    except Exception as e:
        spy_log(f"[AI] 連線異常：{e}")
    return None

async def scrape_moovo():
    """滲透 Moovo 官網抓取數據"""
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page()
        try:
            await page.goto("https://www.ridemoovo.com/city_map_Taipei", wait_until="domcontentloaded", timeout=30000)
            await page.wait_for_selector('.city-location-row', timeout=20000)
            data = await page.evaluate('''() => {
                return Array.from(document.querySelectorAll('.city-location-row')).map(row => {
                    const cells = row.querySelectorAll('.city-location-td');
                    return { name: cells[1].innerText.trim(), bikes: cells[2].innerText.trim() };
                });
            }''')
            await browser.close(); return data
        except Exception as e:
            spy_log(f"[System] 抓取崩潰: {e}"); await browser.close(); return None

def analyze_logic(current_data):
    """核心分析邏輯：數據分類與熱點標註"""
    history = {}
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f: history = json.load(f)
        except: history = {}
    
    changed, unchanged, new_history = [], [], {}
    max_abs_diff = 0
    
    for item in current_data:
        name, curr = item['name'], int(item['bikes'])
        prev_data = history.get(name, {})
        prev = prev_data.get("bikes", curr) if isinstance(prev_data, dict) else int(prev_data)
        
        diff = curr - prev
        abs_diff = abs(diff); max_abs_diff = max(max_abs_diff, abs_diff)
        info = {"name": name, "curr": curr, "diff": f"{'+' if diff > 0 else ''}{diff}", "abs_diff": abs_diff}
        
        if diff != 0: changed.append(info)
        else: unchanged.append(info)
        new_history[name] = {"bikes": curr}

    for item in changed:
        if item['abs_diff'] == max_abs_diff and max_abs_diff > 0: item['hot'] = True
    
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(new_history, f, ensure_ascii=False)
        
    return changed, unchanged

async def main():
    raw_data = await scrape_moovo()
    if raw_data:
        # 1. 數據分類分析
        changed, unchanged = analyze_logic(raw_data)
        
        # 2. Excel 生成與傳輸
        excel_path = create_beautified_excel(changed, unchanged, EXCEL_FILE)
        file_url = upload_excel(excel_path)
        
        # 3. AI 撰寫報告
        prompt = f"""
你現在是專業特工分析官。請撰寫監測報告。
指令：
1. 結構：分兩區「⚡ 變動場站」與「⚪ 穩定場站」。
2. 標註：'hot':True 的站名前加上 🔥。
3. 格式：每站一行，格式：站名:台數 (較上次:變動)。
4. 語言：100% 繁體中文。時間：{current_time_tp}。
數據：變動:{changed}，穩定:{unchanged}
"""
        final_msg = call_gemini(prompt)
        
        # 4. 傳送回報
        if final_msg:
            send_line(f"[🧠 AI 深度情報]\n" + final_msg.strip(), file_url)
        else:
            # 備援分類報告
            report = f"【📡 備援分類報告】時間 {current_time_tp}\n\n"
            report += "⚡ 【變動場站】\n"
            for i in changed:
                p = "🔥" if i.get('hot') else " "
                report += f" {p}{i['name']}:{i['curr']} (較上次:{i['diff']})\n"
            report += "\n⚪ 【穩定場站】\n"
            for i in unchanged:
                report += f"   {i['name']}:{i['curr']} (較上次:0)\n"
            send_line(report, file_url)
    else: spy_log("[System] 偵察失敗")

if __name__ == "__main__":
    asyncio.run(main())
